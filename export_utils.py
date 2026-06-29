"""
Export/Import utilities for MEGA Account Generator.
Supports JSON and Excel formats.
"""
import json
import csv
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


def export_to_json(accounts_data, filepath):
    """
    Export accounts to JSON format.
    
    Args:
        accounts_data: List of account rows from CSV
        filepath: Path to save JSON file
    """
    export_data = {
        "exported_at": datetime.now().isoformat(),
        "version": "1.0",
        "total_accounts": len(accounts_data),
        "accounts": []
    }
    
    for acc in accounts_data:
        account_dict = {
            "email": acc[0] if len(acc) > 0 else "",
            "password": acc[1] if len(acc) > 1 else "",
            "storage_used": acc[2] if len(acc) > 2 else "0 B",
            "free_storage": acc[3] if len(acc) > 3 else "20 GB",
            "session_status": acc[4] if len(acc) > 4 else "Unknown",
            "tags": acc[5].split(',') if len(acc) > 5 and acc[5] else [],
            "mailtm_password": acc[6] if len(acc) > 6 else "",
            "mailtm_id": acc[7] if len(acc) > 7 else ""
        }
        export_data["accounts"].append(account_dict)
    
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(export_data, f, indent=2, ensure_ascii=False)
    
    return len(accounts_data)


def import_from_json(filepath):
    """
    Import accounts from JSON format.
    
    Args:
        filepath: Path to JSON file
        
    Returns:
        List of account rows
    """
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    accounts = []
    for acc in data.get("accounts", []):
        row = [
            acc.get("email", ""),
            acc.get("password", ""),
            acc.get("storage_used", "0 B"),
            acc.get("free_storage", acc.get("total_storage", "20 GB")), # Support both
            acc.get("session_status", "Unknown"),
            ','.join(acc.get("tags", [])),
            acc.get("mailtm_password", ""),
            acc.get("mailtm_id", "")
        ]
        accounts.append(row)
    
    return accounts


def export_to_excel(accounts_data, filepath):
    """
    Export accounts to Excel format with formatting.
    
    Args:
        accounts_data: List of account rows from CSV
        filepath: Path to save Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "MEGA Accounts"
    
    # Headers
    headers = ["Email", "Password", "Storage Used", "Free Storage", 
               "Session Status", "Tags", "Mail.tm Password", "Mail.tm ID"]
    ws.append(headers)
    
    # Format headers
    header_fill = PatternFill(start_color="2CC985", end_color="2CC985", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data
    for acc in accounts_data:
        row_data = list(acc)
        # Ensure row has all columns
        while len(row_data) < 8:
            row_data.append("")
        ws.append(row_data[:8])
    
    # Format status column
    for row_idx in range(2, len(accounts_data) + 2):
        status_cell = ws[f'E{row_idx}']
        if status_cell.value == "Active":
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            status_cell.font = Font(color="006100")
        elif "Failed" in str(status_cell.value):
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            status_cell.font = Font(color="9C0006")
    
    # Adjust column widths
    column_widths = [30, 15, 15, 15, 15, 20, 20, 30]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width
    
    # Add metadata sheet
    meta_ws = wb.create_sheet("Metadata")
    meta_ws.append(["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    meta_ws.append(["Total Accounts", len(accounts_data)])
    meta_ws.append(["Generator", "MEGA Account Generator v2.0"])
    
    wb.save(filepath)
    return len(accounts_data)


def import_from_excel(filepath):
    """
    Import accounts from Excel format.
    
    Args:
        filepath: Path to Excel file
        
    Returns:
        List of account rows
    """
    wb = load_workbook(filepath)
    ws = wb.active
    
    accounts = []
    # Skip header row
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # If email exists
            account = list(row[:8])
            # Ensure all fields exist
            while len(account) < 8:
                account.append("")
            accounts.append(account)
    
    return accounts


def import_from_csv(filepath):
    """
    Import accounts from a plain CSV file.
    Supports formats:
      - email,password
      - email,password,status
      - Full 8-column format matching our CSV
    
    Args:
        filepath: Path to CSV file
        
    Returns:
        List of account rows
    """
    accounts = []
    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
        reader = csv.reader(f)
        for row in reader:
            if not row or not row[0].strip():
                continue
            
            # Skip header rows
            if row[0].strip().lower() in ('email', 'e-mail', 'mail', '#email'):
                continue
            
            email = row[0].strip()
            if '@' not in email:
                continue
            
            password = row[1].strip() if len(row) > 1 else ""
            storage_used = row[2].strip() if len(row) > 2 else "0 B"
            free_storage = row[3].strip() if len(row) > 3 else "20 GB"
            status = row[4].strip() if len(row) > 4 else "Active"
            tags = row[5].strip() if len(row) > 5 else ""
            mailtm_password = row[6].strip() if len(row) > 6 else ""
            mailtm_id = row[7].strip() if len(row) > 7 else ""
            
            accounts.append([
                email, password, storage_used, free_storage,
                status, tags, mailtm_password, mailtm_id
            ])
    
    return accounts


def get_export_stats(accounts_data):
    """
    Get statistics for export summary.
    
    Args:
        accounts_data: List of account rows
        
    Returns:
        dict with stats
    """
    stats = {
        "total": len(accounts_data),
        "active": 0,
        "failed": 0,
        "total_storage_gb": 0,
        "used_storage_gb": 0
    }
    
    for acc in accounts_data:
        if len(acc) > 4:
            if acc[4] == "Active":
                stats["active"] += 1
            elif "Failed" in acc[4]:
                stats["failed"] += 1
        
        # Parse storage (simple approach)
        if len(acc) > 2:
            try:
                used_str = acc[2].replace("GiB", "").replace("GB", "").replace("MiB", "").replace("MB", "").strip()
                if "GiB" in acc[2] or "GB" in acc[2]:
                    stats["used_storage_gb"] += float(used_str.split()[0]) if used_str else 0
            except (ValueError, IndexError):
                pass
    
    stats["total_storage_gb"] = stats["total"] * 20  # Each account has 20GB
    
    return stats
